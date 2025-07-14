import os
import openpyxl
import requests
from requests.exceptions import RequestException
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

# ============================
# Configuration Section
# ============================

# List of Excel files to process - replace these with your own paths
configurations = [
    {'file_path': r"path\\to\\your\\excel_file1.xlsx"},
    {'file_path': r"path\\to\\your\\excel_file2.xlsx"},
    # Add more file paths here as needed
]

# Excel structure assumptions
sheet_indices = [1, 2]         # Only check 2nd and 3rd sheets in each workbook
column_links = 2              # Column B: contains hyperlinks
column_status = 3             # Column C: where status will be written

# Custom user-agent to reduce false bot detections
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36"
}

# Global session reused across all HTTP requests for efficiency
session = requests.Session()

# ============================
# Link Checking Logic
# ============================

def check_link_status(url):
    """Check if a link is reachable (HTTP 200). Uses HEAD for speed, falls back to GET if redirected."""
    try:
        response = session.head(url, headers=headers, allow_redirects=True, timeout=2)

        if response.status_code == 200:
            return "Link is Active"

        if response.is_redirect or response.status_code in [301, 302]:
            final_url = response.headers.get('Location')
            if final_url:
                response = session.get(final_url, headers=headers, timeout=2)
                if response.status_code == 200:
                    return "Link is Active"

        return "Link is Inactive"

    except RequestException:
        return "Link is Inactive"

# ============================
# Excel Processing Logic
# ============================

def process_excel_file(file_path):
    """Open the Excel file, scan defined sheets and update link status in the appropriate column."""
    try:
        with open(file_path, 'a'):
            pass  # Check if file is writable
    except PermissionError:
        print(f"Permission Denied: {file_path} is currently in use. Skipping this file.")
        return

    print(f"Processing file: {file_path}")
    workbook = openpyxl.load_workbook(file_path)

    for sheet_index in sheet_indices:
        sheet = workbook.worksheets[sheet_index]
        print(f"Processing Sheet: {sheet.title}")

        links_to_check = []

        # Loop through rows and collect links
        for row in range(2, sheet.max_row + 1):
            link = sheet.cell(row=row, column=column_links).value
            status = sheet.cell(row=row, column=column_status).value

            if not link:
                continue

            if not status:
                sheet.cell(row=row, column=column_status).value = "Link is Active"
            else:
                links_to_check.append((row, link))

        # Use multithreading for faster processing
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {
                executor.submit(check_link_status, link): row
                for row, link in links_to_check
            }

            for future in tqdm(as_completed(futures), total=len(futures), desc=f"Checking links in {sheet.title}"):
                row = futures[future]
                try:
                    new_status = future.result()
                    sheet.cell(row=row, column=column_status).value = new_status
                except Exception as e:
                    print(f"Error processing row {row}: {e}")

    workbook.save(file_path)
    print(f"Updated file saved: {file_path}")

# ============================
# Entry Point
# ============================

def main():
    if not configurations:
        print("No Excel files found.")
        return

    print(f"Found {len(configurations)} Excel files. Processing...")
    for config in configurations:
        file_path = config['file_path']
        try:
            process_excel_file(file_path)
        except Exception as e:
            print(f"Error processing {file_path}: {e}")

    print("All files processed successfully.")

if __name__ == "__main__":
    main()

file_label.pack(pady=10)

root.mainloop()
